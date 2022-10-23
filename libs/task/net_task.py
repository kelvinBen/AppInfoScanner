#! /usr/bin/python3
# -*- coding: utf-8 -*-
# Author: kelvinBen
# Github: https://github.com/kelvinBen/AppInfoScanner

import openpyxl
import config
from queue import Queue
import libs.core as cores
from libs.core.net import NetThreads


class NetTask(object):
    value_list = []
    domain_list = []

    def __init__(self, result_dict, app_history_list, domain_history_list, file_identifier, threads):
        self.result_dict = result_dict
        self.app_history_list = app_history_list
        self.file_identifier = file_identifier
        self.domain_queue = Queue()
        self.threads = int(threads)
        self.thread_list = []
        self.domain_history_list = domain_history_list

    def start(self):
        xls_result_path = cores.xls_result_path
        workbook = openpyxl.Workbook()
        worksheet = self.__creating_excel_header__(workbook)

        self.__write_result_to_txt__()

        self.__start_threads__(worksheet)

        for thread in self.thread_list:
            thread.join()

        workbook.save(xls_result_path)

    def __creating_excel_header__(self, workbook):
        worksheet = workbook.create_sheet("Result", 0)
        worksheet.cell(row=1, column=1, value="Number")
        worksheet.cell(row=1, column=2, value="IP/URL")
        worksheet.cell(row=1, column=3, value="Domain")
        worksheet.cell(row=1, column=4, value="Status")
        worksheet.cell(row=1, column=5, value="IP")
        worksheet.cell(row=1, column=6, value="Server")
        worksheet.cell(row=1, column=7, value="Title")
        worksheet.cell(row=1, column=8, value="CDN")
        worksheet.cell(row=1, column=9, value="Finger")
        return worksheet

    def __write_result_to_txt__(self):
        append_file_flag = True

        for key, value in self.result_dict.items():
            for result in value:
                if result in self.value_list:
                    continue
                self.value_list.append(result)

                if (("http://" in result) or ("https://" in result)) and ("." in result):
                    if "{" in result or "}" in result or "[" in result or "]" in result or "\\" in result or "!" in result or "," in result:
                        continue

                    domain = result.replace(
                        "https://", "").replace("http://", "")
                    if "/" in domain:
                        domain = domain[:domain.index("/")]

                    if "|" in result:
                        result = result[:result.index("|")]
                    # 目前流通的域名中加上协议头最短长度为11位
                    if len(result) <= 10:
                        continue

                    url_suffix = result[result.rindex(".")+1:].lower()
                    if not(cores.resource_flag and url_suffix in config.sniffer_filter):
                        self.domain_queue.put(
                            {"domain": domain, "url_ip": result})

                    for identifier in self.file_identifier:
                        if identifier in self.app_history_list:
                            if not(domain in self.domain_history_list):
                                self.domain_list.append(domain)
                                self.__write_content_in_file__(
                                    cores.domain_history_path, domain)
                            continue

                        if not(domain in self.domain_list):
                            self.domain_list.append(domain)
                            self.__write_content_in_file__(
                                cores.domain_history_path, domain)

                        if append_file_flag:
                            self.__write_content_in_file__(
                                cores.app_history_path, identifier)
                            append_file_flag = False

    def __start_threads__(self, worksheet):
        for threadID in range(0, self.threads):
            name = "Thread - " + str(threadID)
            thread = NetThreads(threadID, name, self.domain_queue, worksheet)
            thread.start()
            self.thread_list.append(thread)

    def __write_content_in_file__(self, file_path, content):
        with open(file_path, "a+", encoding='utf-8', errors='ignore') as f:
            f.write(content+"\r")
            f.close()
